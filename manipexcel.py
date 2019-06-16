import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import Workbook, load_workbook

A = ['Contas', 'Luz', 'Água', 'Gás', 'Internet']
val = ['Valor', '80', '60', '70', '120']
c = 0

# Escrevendo
# Criando uma planilha
book = Workbook()
sheet = book.active

while c < len(A):
    sheet['A{}'.format(c+1)].value = A[c]
    c += 1

book.save('Tabela.xlsx')

c = 0

# Carregado uma planilha
book = load_workbook('Tabela.xlsx')
sheet = book['Sheet']

while c < len(val):
    sheet['B{}'.format(c+1)].value = val[c]
    c += 1

book.save('Tabela.xlsx')

# Lendo

tabe = pd.read_excel('Tabela.xlsx')

# Mostrar tudo da coluna
#print(tabe['Contas'])

# Linha
for i in tabe.index:
    con = tabe['Contas'][i]
    val = tabe['Valor'][i]
    print('{}: {}'.format(con, val))
